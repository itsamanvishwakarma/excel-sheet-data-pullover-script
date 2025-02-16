import pandas as pd
from openpyxl import load_workbook

# Configuration - Update these values according to your needs
PRIMARY_FILE_PATH = r'C:\Users\itsam\Downloads\primary.xlsx'
PRIMARY_SHEET_NAME = 'Sheet1'
SECONDARY_FILE_PATH = r'C:\Users\itsam\Downloads\secondary.xlsx'
SECONDARY_SHEET_NAME = 'Sheet1'

def update_secondary_sheet():
    # Load data from both sheets
    primary_df = pd.read_excel(
        PRIMARY_FILE_PATH, 
        sheet_name=PRIMARY_SHEET_NAME,
        skiprows=1,
        usecols=[2, 3, 4, 16]  # C=2, D=3, E=4, Q=16
    )
    
    # Add the actual Excel row numbers
    primary_df['excel_row'] = range(3, len(primary_df) + 3)
    
    secondary_df = pd.read_excel(
        SECONDARY_FILE_PATH, 
        sheet_name=SECONDARY_SHEET_NAME,
        skiprows=2,
        usecols=[2, 3, 4, 17]
    )
    
    # Add the actual Excel row numbers for secondary file
    secondary_df['excel_row'] = range(3, len(secondary_df) + 3)
    
    # Rename columns to work with the data
    column_names = {
        primary_df.columns[0]: 'Item Name',    # C column
        primary_df.columns[1]: 'Color',        # D column
        primary_df.columns[2]: 'Opening Qty',  # E column
        primary_df.columns[3]: 'Physical Count' # Q column
    }
    primary_df.rename(columns=column_names, inplace=True)
    secondary_df.rename(columns=column_names, inplace=True)
    
    # Print the first few rows of each dataframe to verify data
    print("Primary DataFrame first few rows:")
    print(primary_df.head())
    print("\nSecondary DataFrame first few rows:")
    print(secondary_df.head())
    
    # Define columns for matching
    match_columns = ['Item Name', 'Color', 'Opening Qty']
    
    # Create a dictionary of primary data for updating
    update_dict = {}
    for _, row in primary_df.iterrows():
        key = tuple(row[match_columns])
        update_dict[key] = {'value': row['Physical Count'], 'row': row['excel_row']}
    
    # Load the existing workbook
    wb = load_workbook(SECONDARY_FILE_PATH)
    ws = wb[SECONDARY_SHEET_NAME]
    
    # Track updates for reporting
    updates_made = 0
    no_matches = 0
    
    # Update Physical Count where matches exist, put 'NA' where no match found
    for _, row in secondary_df.iterrows():
        match_key = tuple(row[match_columns])
        excel_row = row['excel_row']
        
        if match_key in update_dict:
            value = update_dict[match_key]['value']
            ws.cell(row=excel_row, column=17, value=value)
            updates_made += 1
            print(f"Updated row {excel_row}: {match_key} with value {value}")
        else:
            ws.cell(row=excel_row, column=17, value='NA')
            no_matches += 1
            print(f"No match found for row {excel_row}: {match_key} - set to NA")
    
    # Save the workbook
    wb.save(SECONDARY_FILE_PATH)
    print(f"\nTotal updates made: {updates_made}")
    print(f"Total rows with no matches (set to NA): {no_matches}")

if __name__ == '__main__':
    update_secondary_sheet()
    print("\nSecondary sheet updated successfully!")